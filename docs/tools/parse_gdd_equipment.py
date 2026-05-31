"""
Properly parse GDD Equipment sheet.

Each item spans 3 rows:
  Row N+0: No, [Icon], Name, Category, "Stats Category", "Grade skill",
           [grade-skill text per quality: Normal..Epic], [blank], "Max lv Item",
           [max_lvl per quality], ...
  Row N+1: blanks, "Stats base", [base values per quality]
  Row N+2: blanks, "Stats bonus with each level", [bonus per level per quality]

Output: docs/gdd-dump/equipment_parsed.json — list of dicts:
  { name, category, statCategory,
    quality: {
      Normal:  { base, bonus_per_level, max_level, grade_skill_text },
      Good:    {...},
      Better:  {...},
      Excellent: {...},
      Epic:    {...},
    }
  }
"""
import openpyxl, json, sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = 'GDD - GoGo Survival.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['Equipment']

QUALITIES = ['Normal', 'Good', 'Better', 'Excellent', 'Epic']

def cell(row, col):
    v = ws.cell(row=row, column=col).value
    return v

def is_item_header_row(row):
    # column 1 has a numeric No (like 1.0, 2.0, ...), column 3 is name
    no = cell(row, 1)
    name = cell(row, 3)
    if no is None or name is None: return False
    try: float(no)
    except: return False
    return True

def float_or_none(v):
    if v is None: return None
    if isinstance(v, str) and v.strip().lower() in ('n/a', ''): return None
    try: return float(v)
    except: return None

items = []
r = 1
while r <= ws.max_row:
    if not is_item_header_row(r):
        r += 1; continue
    name = str(cell(r, 3)).strip()
    category = cell(r, 4)
    stat_cat = cell(r, 5)
    # grade skill text per quality at columns 7..11 (Normal..Epic)
    grade_skill = {q: cell(r, 7 + i) for i, q in enumerate(QUALITIES)}
    # max lv per quality at columns 14..18 (after Max lv Item header in col 13)
    max_levels = {q: float_or_none(cell(r, 14 + i)) for i, q in enumerate(QUALITIES)}

    # Stats base row (r+1)
    bases = {q: float_or_none(cell(r + 1, 7 + i)) for i, q in enumerate(QUALITIES)}
    # Stats bonus row (r+2)
    bonus = {q: float_or_none(cell(r + 2, 7 + i)) for i, q in enumerate(QUALITIES)}

    quality_data = {}
    for q in QUALITIES:
        quality_data[q] = {
            'base': bases.get(q),
            'bonus_per_level': bonus.get(q),
            'max_level': max_levels.get(q),
            'grade_skill_text': grade_skill.get(q),
        }
    items.append({
        'name': name,
        'category': category,
        'statCategory': stat_cat,
        'quality': quality_data,
    })
    r += 3

os.makedirs('docs/gdd-dump', exist_ok=True)
with open('docs/gdd-dump/equipment_parsed.json', 'w', encoding='utf-8') as f:
    json.dump(items, f, ensure_ascii=False, indent=1)

print(f'Parsed {len(items)} equipment items.')
for it in items[:5]:
    print(f"  {it['name']:<25} ({it['category']}, {it['statCategory']}):")
    for q, d in it['quality'].items():
        print(f"    {q:<10} base={d['base']!s:<8} bonus={d['bonus_per_level']!s:<8} max_lv={d['max_level']!s:<8} skill={d['grade_skill_text']}")
